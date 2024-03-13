"""This function loads an already done analysis.

Input is the folder path.
"""

# from importlib import reload
import numpy as np
import czifile
from openpyxl import load_workbook
from PyQt5 import QtWidgets
from aicsimageio import AICSImage
from skimage.filters import gaussian

import LoadRawData
import SaveReadMatrix
import ServiceWidgets


class RawData:
    """Load and concatenate .czi files."""
    def __init__(self, analysis_folder, fnames):

        first_fr  =  None
        last_fr   =  None
        raw_data  =  None

        # err_msg         =  QtWidgets.QMessageBox()
        first_frame     =  np.load(analysis_folder + '/first_mip_frame.npy')                                               # load the first analyzed raw frame
        last_frame      =  np.load(analysis_folder + '/last_mip_frame.npy')                                                # load the last analyzed raw frame
        ch_numb         =  np.load(analysis_folder + '/ch_numb.npy')                                                       # load channels number
        crop_roi        =  np.load(analysis_folder + '/crop_roi.npy')                                                      # load crop corners coordinates
        pre_nopre_flag  =  np.load(analysis_folder + '/pre_nopre_flag.npy')                                                # load presmoothing / no-presmoothing flag

        pbar  =  ServiceWidgets.ProgressBar(total1=len(fnames))
        pbar.show()
        pbar.update_progressbar1(1)

        for cnt, fname in enumerate(fnames):
            img       =  AICSImage(fname)                                          # read and load file
            pbar.update_progressbar1(cnt)
            raw_data  =  img.get_image_data("TZXY", C=ch_numb)                     # get image
            if pre_nopre_flag:                                                     # if presmoothing is required
                for tt in range(raw_data.shape[0]):                                # frame by frame
                    bff           =  gaussian(raw_data[tt], 1.5)                   # gaussian fitting with a fixed kernel of 1.5
                    raw_data[tt]  =  (bff * 1000).astype(np.uint16)                # since gaussian smoothing gives real values in [0, 1] we just fix this problem

            raw_data_mip  =  np.zeros((img.dims.T, img.dims.X, img.dims.Y), dtype=raw_data.dtype)   # initialize empty matrix for mip
            for uu in range(img.dims.T):
                for xx in range(img.dims.X):
                    raw_data_mip[uu, xx, :]  =  raw_data[uu, :, xx, :].max(0)                               # mip

            raw_data_mip  =  raw_data_mip[:, crop_roi[0]:crop_roi[2], crop_roi[1]:crop_roi[3]]            # crop mip matrix accodringly with crop info
            raw_data      =  raw_data[:, :, crop_roi[0]:crop_roi[2], crop_roi[1]:crop_roi[3]]             # crop raw matrix accodringly with crop info

            uu_first  =  np.sum(np.abs(raw_data_mip - first_frame), axis=(1, 2))                          # subtract 'first_frame' from all the images of the mip stack and then sum in x-y to find the first analyzed frame
            if uu_first.min() == 0:                                                                       # if the min is equal to 0, the first analyzed frame is in the loaded data, otherwise go on with the other file overwriting
                first_fr      =  np.where(uu_first == 0)[0][0]                                            # check the frame number corresponding to the first analyzed frame
                raw_data      =  raw_data[first_fr:]                                                      # cut properly 4d raw data
                raw_data_mip  =  raw_data_mip[first_fr:]                                                  # cut properly mip raw data
                fname2st      =  cnt                                                                      # store the position in the fnames list of the file
                break                                                                                     # get out of the for loop

        if np.where(np.sum(np.abs(raw_data_mip - last_frame), axis=(1, 2)) == 0)[0].size == 0:          # check if the last analyzed frame is in the same file: if not go on loading the following files
            for ff in fnames[fname2st + 1:]:                                                            # starting from the file following the one with the first analyzed frame
                pbar.update_progressbar1(cnt + 1)
                cnt          +=  1
                img           =  AICSImage(ff)                                                          # read and load file
                raw_data_bff  =  img.get_image_data("TZXY", C=ch_numb)                                  # get image
                if pre_nopre_flag:                                                                      # presmooth if needed
                    for tt in range(raw_data_bff.shape[0]):
                        bff               =  gaussian(raw_data_bff[tt], 1.5)
                        raw_data_bff[tt]  =  (bff * 1000).astype(np.uint16)
                raw_data_mip_bff  =  np.zeros((img.dims.T, img.dims.X, img.dims.Y), dtype=raw_data_bff.dtype)   # initialize empty matrix for mip
                for uu in range(img.dims.T):
                    for xx in range(img.dims.X):
                        raw_data_mip_bff[uu, xx, :]  =  raw_data_bff[uu, :, xx, :].max(0)                               # mip

                raw_data_mip  =  raw_data_mip[:, crop_roi[0]:crop_roi[2], crop_roi[1]:crop_roi[3]]            # crop mip matrix accodringly with crop info
                raw_data      =  raw_data[:, :, crop_roi[0]:crop_roi[2], crop_roi[1]:crop_roi[3]]             # crop raw matrix accodringly with crop info

                uu_last  =  np.sum(np.abs(raw_data_mip_bff - last_frame), axis=(1, 2))                        # subtract 'last_frame' from all the images of the mip stack and then sum in x-y to find the first analyzed frame
                if uu_last.min() == 0:                                                                        # if the last analyzed frame is in the loaded file
                    last_fr           =  np.where(uu_last == 0)[0][0]                                         # search its coordinate
                    raw_data_bff      =  raw_data_bff[:last_fr + 1]                                           # properly cut the 4d raw data
                    raw_data_mip_bff  =  raw_data_mip_bff[:last_fr + 1]                                       # properly cut the mip raw data
                    raw_data          =  np.concatenate((raw_data, raw_data_bff), axis=0)                     # concatenate 4d
                    raw_data_mip      =  np.concatenate((raw_data_mip, raw_data_mip_bff), axis=0)             # concatenate mip
                    break                                                                                     # get out of the loop
                else:                                                                                         # if the last analyzed frame is not there, just concatenate and go on
                    raw_data      =  np.concatenate((raw_data, raw_data_bff), axis=0)
                    raw_data_mip  =  np.concatenate((raw_data_mip, raw_data_mip_bff), axis=0)

        elif np.where(np.sum(np.abs(raw_data_mip - last_frame) == 0))[0].size != 0:                         # if the last analyzed frame is in the same file
            uu_last           =  np.sum(np.abs(raw_data_mip - last_frame), axis=(1, 2))                     # find the coordinate of the last analyzed frame
            last_fr           =  np.where(uu_last == 0)[0][0]
            raw_data_bff      =  raw_data[:last_fr + 1]                                                     # cut properly the 4d raw data
            raw_data_mip_bff  =  raw_data_mip[:last_fr + 1]                                                 # cut properly the mip raw data

        try:
            pix_size_xy  =  img.physical_pixel_sizes.X                                                      # read pixel size
            pix_size_z   =  img.physical_pixel_sizes.Z

        except:
            pix_size_xy, pix_size_z  =  ServiceWidgets.InputPixSize.get_vals()                              # if for any reason it is not possible to read pixel size, launch a popup tool to introduce it manually

        with czifile.CziFile(str(fnames[0])) as czi:                                                # read the time step
            for attachment in czi.attachments():
                if attachment.attachment_entry.name == 'TimeStamps':
                    timestamps  =  attachment.data()
                    break
            else:
                raise ValueError('TimeStamps not found')
        time_step_value  =  np.round(timestamps[1] - timestamps[0], 2)                    # time step value

        self.raw_data         =  raw_data
        self.raw_data_mip     =  raw_data_mip
        self.pix_size_z       =  pix_size_z
        self.pix_size_xy      =  pix_size_xy
        self.ch_numb          =  ch_numb
        self.time_step_value  =  time_step_value


class AnalysisParameters:
    """Load analysis parameters."""
    def __init__(self, analysis_folder):

        wb                       =  load_workbook(analysis_folder + '/SpotsAnalysis.xlsx')
        sm_sheet                 =  wb[wb.sheetnames[0]]
        self.merge_radius_value  =  int(sm_sheet.cell(column=2, row=1).value)
        self.spts_thr_value      =  sm_sheet.cell(column=2, row=2).value
        self.min_volthr_value    =  int(sm_sheet.cell(column=2, row=3).value)
        self.dist_thr_value      =  int(sm_sheet.cell(column=2, row=4).value)


class SpotsTracked:
    """Load tracked spots."""
    def __init__(self, analysis_folder):

        self.spts_trck_fin  =  SaveReadMatrix.SpotsMatrixReader(analysis_folder, '/spots_trck.npy').spts_lbls
        self.orig_tags      =  np.load(analysis_folder + '/spots_orig_tags.npy')


class SpotsDetected:
    """Load detected spots."""
    def __init__(self, analysis_folder):

        self.spots_ints     =  SaveReadMatrix.SpotsMatrixReader(analysis_folder, '/spots_ints.npy').spts_lbls
        self.spots_vol      =  SaveReadMatrix.SpotsMatrixReader(analysis_folder, '/spots_vol.npy').spts_lbls
        self.spots_2d_lbls  =  SaveReadMatrix.SpotsMatrixReader(analysis_folder, '/spots_2d_lbls.npy').spts_lbls
        self.spots_coords   =  np.load(analysis_folder + '/spots_coords.npy')
