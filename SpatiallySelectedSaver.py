"""This function writes the results of a spatial analysis.

Input are the analysis folder and the xlsx file name to write results.
"""


import datetime
import numpy as np
import xlsxwriter
from openpyxl import load_workbook


class SpatiallySelectedSaver:
    """Onmy class, does all the job."""
    def __init__(self, xlsx_filename, analysis_folder, roi_letf_value, roi_right_value, roi_top_value, roi_bottom_value, pix_size_xy):

        spots_features        =  np.load(analysis_folder + '/spots_features.npy')                                       # load spots features from analysis results
        spots_idxs            =  np.load(analysis_folder + '/spots_idxs.npy')                                           # load the array of spots' tags
        spots_features_slctd  =  np.zeros_like(spots_features)                                                          # initialize the matrix of the selected spots
        spots_idxs_slctd      =  []                                                                                     # initialize the list of the selected spots tags

        cnt  =  0
        for uu in range(spots_features.shape[2]):
            ref_x  =  spots_features[:, 5, uu]                                                                          # read x coordinates of each spot all along the evolution
            ref_x  =  ref_x[ref_x != 0]                                                                                 # remove frames in which it is abbsent (in this case it is marked as 0 and this fakes the average)
            ref_y  =  spots_features[:, 6, uu]                                                                          # same as before but in y
            ref_y  =  ref_y[ref_y != 0]
            if roi_letf_value < ref_x.mean() <= roi_right_value and roi_top_value < ref_y.mean() <= roi_bottom_value:   # check if the average centroid position is inside the square of the ROI
                spots_features_slctd[:, :, cnt]  =  spots_features[:, :, uu]                                            # if it is the case, add the spot to the list of the spots inside
                spots_idxs_slctd.append(spots_idxs[uu])
                cnt  +=  1

        spots_features_slctd  =  spots_features_slctd[:, :, :cnt]                                                       # cut the array removing the empty rows

        wb  =  load_workbook(analysis_folder + '/SpotsAnalysis.xlsx')                                                   # write results
        sh  =  wb[wb.sheetnames[0]]

        if xlsx_filename[-5:] != '.xlsx':
            xlsx_filename += '.xlsx'

        book    =  xlsxwriter.Workbook(xlsx_filename)
        sheet1  =  book.add_worksheet("Info")
        sheet2  =  book.add_worksheet("Volume")
        sheet3  =  book.add_worksheet("Intensity")
        sheet4  =  book.add_worksheet("Background")
        sheet5  =  book.add_worksheet("Ints by Bckg")
        sheet6  =  book.add_worksheet("Coords")

        sheet1.write(0, 0, "merge radius")
        sheet1.write(0, 1, sh["B1"].value)
        sheet1.write(1, 0, "detect threshold")
        sheet1.write(1, 1, sh["B2"].value)
        sheet1.write(2, 0, "min volume thr")
        sheet1.write(2, 1, sh["B3"].value)
        sheet1.write(3, 0, "Spots dist Thr")
        sheet1.write(3, 1, sh["B4"].value)
        sheet1.write(4, 0, "Min number of active frames")
        sheet1.write(4, 1, sh["B5"].value)
        sheet1.write(5, 0, "pixels size x-y (µm)")
        sheet1.write(5, 1, sh["B6"].value)
        sheet1.write(6, 0, "pixels size z (µm)")
        sheet1.write(6, 1, sh["B7"].value)
        sheet1.write(7, 0, "time step (s)")
        sheet1.write(7, 1, sh["B8"].value)
        sheet1.write(8, 0, "left")
        sheet1.write(8, 1, np.round(roi_letf_value * pix_size_xy, 2))
        # sheet1.write(8, 1, np.round(roi_letf_value * pix_size_xy * 1000000, 2))
        sheet1.write(9, 0, "right")
        sheet1.write(9, 1, np.round(roi_right_value * pix_size_xy, 2))
        # sheet1.write(9, 1, np.round(roi_right_value * pix_size_xy * 1000000, 2))
        sheet1.write(10, 0, "top")
        sheet1.write(10, 1, np.round(roi_top_value * pix_size_xy, 2))
        # sheet1.write(10, 1, np.round(roi_top_value * pix_size_xy * 1000000, 2))
        sheet1.write(11, 0, "bottom")
        sheet1.write(11, 1, np.round(roi_bottom_value * pix_size_xy, 2))
        # sheet1.write(11, 1, np.round(roi_bottom_value * pix_size_xy * 1000000, 2))

        sheet1.write(12, 0, "software version")
        sheet1.write(12, 1, sh["B10"].value)
        sheet1.write(13, 0, "date")
        sheet1.write(13, 1, datetime.datetime.now().strftime("%d-%b-%Y"))
        sheet1.write(15, 0, "File names")
        for ll in range(14, sh.max_row + 1):
            sheet1.write(16 + ll - 14, 0, sh.cell(row=ll, column=1).value)

        sheet2.write(0, 0, "Time")
        sheet3.write(0, 0, "Time")
        sheet4.write(0, 0, "Time")
        sheet5.write(0, 0, "Time")
        sheet6.write(0, 0, "Time")
        for k in range(spots_features_slctd.shape[0]):
            sheet2.write(k + 1, 0, k)
            sheet3.write(k + 1, 0, k)
            sheet4.write(k + 1, 0, k)
            sheet5.write(k + 1, 0, k)
            sheet6.write(k + 2, 0, k)

        if len(spots_idxs_slctd) == 0:
            for ccnt, spot_idx in enumerate(spots_idxs_slctd):
                sheet2.write(0, 1 + ccnt, "Spts_" + str(spot_idx))
                sheet3.write(0, 1 + ccnt, "Spts_" + str(spot_idx))
                sheet4.write(0, 1 + ccnt, "Spts_" + str(spot_idx))
                sheet5.write(0, 1 + ccnt, "Spts_" + str(spot_idx))
                sheet6.write(0, 1 + ccnt * 3, "Spts_" + str(spot_idx))
                sheet6.write(1, 1 + ccnt * 3, "z-coord")
                sheet6.write(1, 2 + ccnt * 3, "x-coord")
                sheet6.write(1, 3 + ccnt * 3, "y-coord")

        elif len(spots_idxs_slctd) != 0:
            for ccnt, spot_idx in enumerate(spots_idxs_slctd):
                sheet2.write(0, 1 + ccnt, "Spts_" + str(spot_idx))
                sheet3.write(0, 1 + ccnt, "Spts_" + str(spot_idx))
                sheet4.write(0, 1 + ccnt, "Spts_" + str(spot_idx))
                sheet5.write(0, 1 + ccnt, "Spts_" + str(spot_idx))
                sheet6.write(0, 1 + ccnt * 3, "Spts_" + str(spot_idx))
                sheet6.write(1, 1 + ccnt * 3, "z-coord")
                sheet6.write(1, 2 + ccnt * 3, "x-coord")
                sheet6.write(1, 3 + ccnt * 3, "y-coord")

        for ii1 in range(spots_features_slctd.shape[0]):
            for ii2 in range(spots_features_slctd.shape[2]):
                sheet2.write(1 + ii1, 1 + ii2, spots_features_slctd[ii1, 0, ii2])
                sheet3.write(1 + ii1, 1 + ii2, spots_features_slctd[ii1, 1, ii2])
                sheet4.write(1 + ii1, 1 + ii2, spots_features_slctd[ii1, 2, ii2])
                sheet5.write(1 + ii1, 1 + ii2, spots_features_slctd[ii1, 3, ii2])
                sheet6.write(2 + ii1, 1 + ii2 * 3, spots_features_slctd[ii1, 4, ii2])
                sheet6.write(2 + ii1, 2 + ii2 * 3, spots_features_slctd[ii1, 5, ii2])
                sheet6.write(2 + ii1, 3 + ii2 * 3, spots_features_slctd[ii1, 6, ii2])

        book.close()
