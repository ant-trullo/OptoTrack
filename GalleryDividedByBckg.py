"""This function plots the gallery of spots intensity time traces
divided by background.

Input is the xlsx file of the analysis.

"""


# import time
import numpy as np
import pyqtgraph as pg
from openpyxl import load_workbook


class GalleryDividedByBckg:
    """Only class, does all the job."""
    def __init__(self, xlsx_filename):

        workbook    =  load_workbook(xlsx_filename)                                                                     # load the xlsx file
        sheet_bckg  =  workbook[workbook.sheetnames[4]]                                                                 # read the proper sheet

        spts_vals  =  np.zeros((sheet_bckg.max_column, sheet_bckg.max_row))                                             # initialize the matrix: spots tag in row and intensity in column
        for uu in range(1, sheet_bckg.max_row):
            for vv in range(1, sheet_bckg.max_column):
                spts_vals[vv - 1, uu - 1]  =  sheet_bckg.cell(row=uu + 1, column=vv + 1).value                          # fill the matrix

        spts_id  =  []                                                                                                  # list of spots tags
        for kk in range(1, sheet_bckg.max_column):
            spts_id.append(int(sheet_bckg.cell(column=kk + 1, row=1).value[5:]))                                        # populate the list of spots tags

        y_sup    =  spts_vals.max()                                                                                     # maximum value in y
        n_rows   =  6
        n_cols   =  7
        num_win  =  len(spts_id) // (n_cols * n_rows) + 1                                                               # number of plot-frames needed
        for win_idxs in range(num_win):                                                                                 # for each plot-frame plot the time series with tag and whatever
            # time.sleep(3)
            # str_win1  =  "win" + str(win_idxs) + "  =  pg.GraphicsWindow()"
            str_win1  =  "win" + str(win_idxs) + "  =  pg.GraphicsLayoutWidget()"
            str_win2  =  "win" + str(win_idxs) + ".setWindowTitle('Transcriptional Traces " + str(win_idxs + 1) + "')"
            str_win3  =  "win" + str(win_idxs) + ".showMaximized()"
            str_win4  =  "win" + str(win_idxs) + ".setBackground('w')"
            exec(str_win1)
            exec(str_win2)
            exec(str_win3)
            exec(str_win4)

            idx_name = 0
            for idx_r in range(n_rows):
                for idx_c in range(n_cols):
                    str_addplot  =  "p" + str(idx_name) +  "= win" + str(win_idxs) + ".addPlot(row=" + str(idx_r) + ", col=" + str(idx_c) + ")"
                    exec(str_addplot)
                    idx_name  +=  1

            for k in range(n_cols * n_rows):
                if k + win_idxs * n_cols * n_rows <= len(spts_id) - 1:
                    str_cmnd1  =  "p" + str(k) + ".plot(spts_vals[k  + win_idxs * n_cols * n_rows, :], pen=pg.mkPen('r', width=4), symbol='o', symbolSize=5)"
                    str_cmnd2  =  "p" + str(k) + ".setYRange(0, y_sup)"
                    str_cmnd3  =  "tag_text" + str(k) + " = pg.TextItem('tag = " + str(spts_id[k + win_idxs * n_cols * n_rows]) + "', color='k')"
                    str_cmnd4  =  "tag_text" + str(k) + ".setPos(1, y_sup)"
                    str_cmnd5  =  "p" + str(k) + ".addItem(tag_text" + str(k) + ")"
                    exec(str_cmnd1)
                    exec(str_cmnd2)
                    exec(str_cmnd3)
                    exec(str_cmnd4)
                    exec(str_cmnd5)
                else:
                    break

        print(StrangePatch)                                                                                             # here I force a bug, otherwise the popup does not work: mystery
